import datetime
import os.path
import re
from tkinter import ttk
from bs4 import BeautifulSoup
import requests
from tkinter import *
import openpyxl

def has_number(text):
    return any(char.isdigit() for char in text)


if __name__ == '__main__':
    root = Tk()
    root.title('Парсер для Krisha.kz')
    root.geometry('960x600')
    root.resizable(False, False)
    root.iconbitmap(default='template/logo.png')

    icon = PhotoImage(file='template/logo.png')
    root.iconphoto(False, icon)

    # * СПИСОК ГОРОДА
    cities = ['Астана']

    label_city = ttk.Label()
    label_city['text'] = 'Выберите город'
    label_city.pack(anchor=NW, padx=6, pady=6)

    combobox_cities = ttk.Combobox(values=cities)
    combobox_cities.pack(anchor=NW, padx=6, pady=6)

    # * СПИСОК РАЙОНЫ
    label_district = ttk.Label()
    label_district['text'] = 'Выберите район'
    label_district.pack(anchor=NW, padx=6, pady=6)

    combobox_districts = ttk.Combobox()
    combobox_districts.pack(anchor=NW, padx=6, pady=6)


    # * РАЙОНЫ В ЗАВИСИМОСТИ ОТ ГОРОДА
    def selected_city(event):
        districts_empty = []
        districts_astana = ['Есильский р-н', 'Алматы р-н', 'Сарыарка р-н', 'р-н Байконур']

        selection = combobox_cities.get()

        if selection == 'Астана':
            combobox_districts.config(values=districts_astana)
        else:
            combobox_districts.config(values=districts_empty)


    combobox_cities.bind("<<ComboboxSelected>>", selected_city)

    # * ФЛАЖОК НЕ ПОСЛЕДНИЙ ЭТАЖ
    last_floor_var = BooleanVar()
    last_floor_checkbutton = ttk.Checkbutton(text='Не последний этаж', variable=last_floor_var)
    last_floor_checkbutton.pack(padx=6, pady=6, anchor=NW)

    # * ФЛАЖОК НЕ ПЕРВЫЙ ЭТАЖ
    first_floor_var = BooleanVar()
    first_floor_checkbutton = ttk.Checkbutton(text='Не первый этаж', variable=first_floor_var)
    first_floor_checkbutton.pack(padx=6, pady=6, anchor=NW)

    # * ФЛАЖОК ЕСТЬ ФОТО
    has_image_var = BooleanVar()
    has_image_checkbutton = ttk.Checkbutton(text='Есть фото', variable=has_image_var)
    has_image_checkbutton.pack(padx=6, pady=6, anchor=NW)

    # * ФЛАЖОК НОВОСТРОЙКИ
    is_new_var = BooleanVar()
    is_new_checkbutton = ttk.Checkbutton(text='Новостройки', variable=is_new_var)
    is_new_checkbutton.pack(padx=6, pady=6, anchor=NW)


    # * КОМНАТЫ
    def selected_rooms(event):
        # получаем индексы выделенных элементов
        selected_indices = rooms_listbox.curselection()
        # получаем сами выделенные элементы
        selected_rooms = ",".join([rooms_listbox.get(i) for i in selected_indices])
        msg = f"Количество комнат: {selected_rooms}"
        selection_label["text"] = msg


    rooms = ['1', '2', '3', '4', '5']
    room_var = Variable(value=rooms)

    selection_label = ttk.Label()
    selection_label.pack(anchor=NW, fill=X, padx=5, pady=5)

    rooms_listbox = Listbox(listvariable=room_var, selectmode=EXTENDED)
    rooms_listbox.pack(anchor=NW, fill=X, padx=5, pady=5)
    rooms_listbox.bind("<<ListboxSelect>>", selected_rooms)

    # * ПУТЬ ДО ПАПКИ
    # def is_valid(new_val):
    #     result = os.path.exists(new_val)
    #
    #     if not result:
    #         errmsg.set("Папки не существует!!!")
    #     else:
    #         errmsg.set("")
    #     return result
    #
    # check = (root.register(is_valid), "%P")
    #
    # errmsg = StringVar()

    # path_entry = ttk.Entry(validate="key", validatecommand=check)
    # path_entry.pack(padx=5, pady=5, anchor=NW)

    # file_label = ttk.Label()
    # file_label['text'] = 'Введите путь до папки где вы хотите сохранить результат работы парсера:'
    # file_label.pack(padx=5, pady=5, anchor=NW)

    path_entry = ttk.Entry()
    path_entry.pack(padx=5, pady=5, anchor=NW)


    # error_label = ttk.Label(foreground="red", textvariable=errmsg, wraplength=250)
    # error_label.pack(padx=5, pady=5, anchor=NW)

    # * КНОПКА СПАРСИТЬ
    def click_button():
        combobox_cities_variable = combobox_cities.get()
        combobox_districts_variable = combobox_districts.get()

        # flag_last_floor = last_floor_var.get()
        # flag_first_floor = first_floor_var.get()
        # flag_has_image = has_image_var.get()
        # flag_is_new = is_new_var.get()

        selected_indices = rooms_listbox.curselection()
        lst_rooms = [rooms_listbox.get(i) for i in selected_indices]

        def generate_url():
            BASE_URL = "https://krisha.kz/prodazha/kvartiry/"

            if combobox_cities_variable == 'Астана':
                district_mapping = {
                    'Есильский р-н': 'astana-esilskij',
                    'Алматы р-н': 'astana-almatinskij',
                    'Сарыарка р-н': 'astana-saryarkinskij',
                    'р-н Байконур': 'r-n-bajkonur'
                }
                BASE_URL = BASE_URL + district_mapping[combobox_districts_variable] + '/?'
            if has_image_var.get() is True:
                BASE_URL = BASE_URL + 'das[_sys.hasphoto]=1'
            if first_floor_var.get() is True:
                BASE_URL = BASE_URL + '&das[floor_not_first]=1'
            if last_floor_var.get() is True:
                BASE_URL = BASE_URL + '&das[floor_not_last]=1'
            if len(lst_rooms):
                for room in lst_rooms:
                    BASE_URL = BASE_URL + '&das[live.rooms][]=' + str(room)
            if is_new_var.get() is True:
                BASE_URL = BASE_URL + '&das[novostroiki]=1'

            print(BASE_URL)
            return BASE_URL

        BASE_URL = generate_url()
        page = requests.get(BASE_URL)
        soup = BeautifulSoup(page.content, "html.parser")

        pages_lst = soup.find_all("a", class_="paginator__btn")

        max_pages = ''

        for page in pages_lst:
            text = page.text.replace(' ', '').replace('\n', '')
            if has_number(text):
                max_pages = text

        index_text = 2
        index_price = 2

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 'Объявление'
        ws['B1'] = 'Ссылка'
        ws['C1'] = 'Цена'

        if len(pages_lst):
            for i in range(int(max_pages)):
                try:
                    BASE_URL = BASE_URL + '&page=' + str(i + 1)
                    page = requests.get(BASE_URL)
                    soup = BeautifulSoup(page.content, "html.parser")
                except Exception as e_:
                    print(e_)
                    BASE_URL = BASE_URL + '?page=' + str(i + 1)
                    page = requests.get(BASE_URL)
                    soup = BeautifulSoup(page.content, "html.parser")

                flats_hrefs = soup.find_all("div", class_="a-card__header-left")
                for href in flats_hrefs:
                    link_text = href.findNext("a").text
                    link_to_parse = href.findNext("a").get("href")

                    ws['A' + str(index_text)] = link_text
                    ws['B' + str(index_text)] = 'https://krisha.kz' + link_to_parse

                    index_text += 1

                flats_prices = soup.find_all("div", class_="a-card__price")
                for flat_price in flats_prices:
                    price = flat_price.text.replace(' ', '').replace('\n', '')

                    ws['C' + str(index_price)] = price

                    index_price += 1

                today = datetime.datetime.today().date()
                wb.save('Результат ' + str(today) + '.xlsx')

        else:
            page = requests.get(BASE_URL)
            soup = BeautifulSoup(page.content, "html.parser")

            flats_hrefs = soup.find_all("div", class_="a-card__header-left")
            for href in flats_hrefs:
                link_text = href.findNext("a").text
                link_to_parse = href.findNext("a").get("href")

                ws['A' + str(index_text)] = link_text
                ws['B' + str(index_text)] = 'https://krisha.kz' + link_to_parse

                index_text += 1

            flats_prices = soup.find_all("div", class_="a-card__price")
            for flat_price in flats_prices:
                price = flat_price.text.replace(' ', '').replace('\n', '')

                ws['C' + str(index_price)] = price

                index_price += 1

            today = datetime.datetime.today().date()
            wb.save('Результат ' + str(today) + '.xlsx')


    btn = ttk.Button(text="Спарсить", command=click_button, width=25)
    btn.pack()

    root.mainloop()
